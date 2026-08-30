#!/usr/bin/env python3
"""
Dubai Data Scientist / AI Engineer job tracker for Apoorv Parmar.

Pulls fresh LinkedIn postings via Apify, scores them against Apoorv's resume
profile, excludes known recruiting-mill / non-fit sources, diffs against
previously-seen job IDs (data/seen_jobs.json), and appends only new rows to
the tracking spreadsheet (data/job_matches.xlsx). Designed to run unattended
on a GitHub Actions cron schedule.

Environment variables required:
  APIFY_TOKEN - Apify API token (repo secret)
"""

import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from pydantic import BaseModel, Field
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.exceptions import ModelRateLimitError

load_dotenv()

# Harmless one-time notice from the google-genai SDK ("use Chat.send_message
# instead of Models.generate_content for AFC") -- not something we control
# since langchain_google_genai calls generate_content internally.
logging.getLogger("google_genai.models").setLevel(logging.ERROR)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

APIFY_TOKEN = os.environ.get("APIFY_TOKEN")
if not APIFY_TOKEN:
    print("ERROR: APIFY_TOKEN environment variable not set.", file=sys.stderr)
    sys.exit(1)

ACTOR_ID = "curious_coder~linkedin-jobs-scraper"
APIFY_RUN_URL = f"https://api.apify.com/v2/acts/{ACTOR_ID}/run-sync-get-dataset-items"

DATA_DIR = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
SEEN_JOBS_PATH = DATA_DIR / "seen_jobs.json"
OUTPUT_XLSX = DATA_DIR / "job_matches.xlsx"

SEARCHES = [
    {"keywords": "Data Scientist", "location": "Germany"},
    {"keywords": "Data Scientist", "location": "Dubai"},
    {"keywords": "Data Scientist", "location": "Netherlands"},
    {"keywords": "Data Scientist", "location": "Abu Dhabi"},
    {"keywords": "Data Scientist", "location": "Denmark"}
]

ATS_THRESHOLD = 75

# ---------------------------------------------------------------------------
# Exclusion rules (from conversation history: recruiting-mill / non-fit sources)
# ---------------------------------------------------------------------------
# These are company-name substrings (case-insensitive) that were identified
# as recruiting-mill noise or otherwise not worth tracking, established in
# the original manual review with Apoorv.

EXCLUDED_COMPANY_SUBSTRINGS = [
    "jobs ai",           # generic AI-recruiting-mill reposts
    "jack & jill",       # AI-recruiter placement gimmick, anonymized client
    "jack and jill",
]

# Postings whose title/description signal they are not a fit regardless of
# keyword match (fresh-grad-only academic posts, pure PhD research positions
# unrelated to industry seniority level).
EXCLUDED_TITLE_PATTERNS = [
    r"\bphd student\b",
    r"\bresearch associate\b.*\bphd\b",
    r"\bwerkstudent\b",       # student/intern roles, not fit for 10+ yr profile
    r"\bpraktikum\b",         # internship
    r"\bintern(ship)?\b",
]

# Salary bands that are clearly below market for a Lead-level 10+ yr profile
# (used as a soft signal only, does not hard-exclude)
LOW_SALARY_PATTERN = re.compile(r"\$\s?(\d{2})[kK]\s?-\s?\$?\s?(\d{2,3})[kK]")


def is_excluded(job: dict) -> bool:
    company = (job.get("companyName") or "").lower()
    title = (job.get("title") or "").lower()

    for substr in EXCLUDED_COMPANY_SUBSTRINGS:
        if substr in company:
            return True

    for pattern in EXCLUDED_TITLE_PATTERNS:
        if re.search(pattern, title, re.IGNORECASE):
            return True

    return False


# ---------------------------------------------------------------------------
# Resume profile (Apoorv Parmar) - used for scoring
# ---------------------------------------------------------------------------

RESUME_PROFILE = """
Apoorv Parmar - Lead Data Scientist, People Analytics, Gartner
10+ years experience across Gartner, Genpact, McKinsey & Company, Accenture, Novozymes, IHS Markit.

Core specialization: Generative AI / LLMs / Agentic AI, NLP. Hands-on with LangChain, LangGraph,
Retrieval Augmented Generation (RAG), Azure OpenAI, multi-agent orchestration, prompt engineering,
vector search (Azure AI Search, hybrid + semantic ranking), LLM-as-a-judge evaluation, RAGAS
framework. Built an agentic attrition-forecasting system at Gartner combining Random Forest with a
LangGraph multi-agent pipeline and RAG over unstructured HR data. Built an AI-powered candidate
scoring pipeline at Gartner using Azure OpenAI for feature extraction at 1,000+ CVs per requisition.
At Genpact, built RAG applications for large US/UK banks using LangChain, Microsoft Phi-3, OpenAI
models, Azure AI Search.

Also has classical ML background: Random Forest, XGBoost, LightGBM, fraud detection, churn
prediction, sentiment analysis (BERT), regression/clustering -- but this is supporting experience,
not his current differentiator.

Tech stack: Python, SQL, Databricks, Azure Data Lake Storage, Azure OpenAI, LangGraph, LangChain,
MLflow, ChromaDB, FAISS, Pinecone.

Extensive expeerience in Finance, Fintech and Banking Domain.

Education: MBA (FORE School of Management) in Finance and IT, B.Tech Computer Science.
Location: Noida, India (applying to roles in Germany; no German fluency stated on his resume).
Seniority: Total 10 years of Experience, 8+ years of experience in Data Science/Gen AI -- roles requiring far more (15+ yrs, Director/VP/Chief-level titles far above
Lead Data Scientist) or far less (internship/Werkstudent/fresh-grad) are weaker seniority fits.
"""

WORK_EXPERIENCE = """Led development of an employee attrition forecasting system using Random Forest on longitudinal
workforce data to proactively identify high-risk employee churn. Designed the solution to predict
future-quarter attrition trends from historical employee behaviour patterns, improving early risk detection
accuracy. Improved model precision by 12% and recall by 5% through temporal feature engineering.
Enabled HR teams to uncover key drivers of attrition, prioritise retention interventions, and support
workforce planning using predictive analytics.

Orchestrated a multi-agent workflow on top of the attrition model using LangGraph and Databricks
Workflows, with monthly batch scoring triggering agent pipelines for all employees breaching the risk
threshold. Implementing a Retrieval Augmented Generation (RAG) component using OpenAI embeddings,
Azure AI Search (Hybrid Search with Semantic Ranking) to index unstructured HR data (exit surveys,
internal job postings, past retention case studies) for recommendation generation. Maintained strict
enterprise guardrails via Azure AI Content Safety and PII masking to protect sensitive employee records,
evaluating the entire ecosystem via automated LLM-as-a-judge pipelines, RAGAS framework on
Groundedness, Context Relevance, Answer Relevance, MMR, Precision@k, recall@k metrics.

AI-Powered Candidate Scoring Pipeline: Designed and implemented an end-to-end candidate evaluation
pipeline supporting high-volume sales hiring (1,000+ CVs per requisition across multiple concurrent
openings). Built scalable data ingestion and feature engineering pipelines using Azure Data Lake Storage
and Azure Databricks, leveraging Azure OpenAI to extract and score candidate skills, business impact,
domain experience, and job-fit signals. Combined LLM generated features with structured workforce data to
predict first-year employee performance and improve hiring decision quality.

Findings & Strategic Outcome: Model achieved 65% precision and recall; Conducted model validation,
performance diagnostics, and stakeholder reviews to identify limitations in historical performance-rating
data used as prediction targets. Presented findings and recommendations to HR leadership, influencing
roadmap decisions and defining a Phase-2 strategy incorporating AI-driven interview assessment signals to
improve predictive accuracy and talent selection outcomes

Developed and deployed generative AI workflow using Langchain, Azure OpenAI API on Customer
feedback data of Credit Card division of a Large US Bank. Large Language Model extracted key information
from the feedback data, collated the information and delivered actionable insights that resulted in a 5%
improvement in NPS for the bank.

Retrieval Augmented Generation (RAG) Implementations: Built a Retrieval Augmented Generation
(RAG) application for a Large UK Bank in their credit card division, leveraging LangChain, Python on 20+
PDF documents, utilising recursive chunking strategy, semantic search retrieval and generating the
response using Open Source Microsoft Phi-3 Model. Used Evaluation strategies like Precision@k,
Recall@k, time to first token, token per sec, BLEU score and LLM as a judge to measure the performance
of the application. The Application reduced the customer hold time by 15% while enhancing credit card
dispute resolution efficiency by 7%. Deployed the application on Azure cloud along with a fully local
deployment.

Developed and deployed a Retrieval Augmented Generation (RAG) application for a Large US Bank for
their Personal Banking division, leveraging LangChain, Python on 50+ PDF documents, utilising recursive
chunking strategy, OpenAI api for embedding, semantic search retrieval on Azure AI search and
generating the response using OpenAI Models. Used Evaluation strategies like Precision@k,
Recall@k,MMR, BLEU score and LLM as a judge along with Langsmith framework to measure the
performance of the application.

Fraud Detection & Predictive Modelling: Designed and deployed scalable AI/ML solutions for Credit Card
Disputes leveraging the Random Forest algorithm, achieving 87% accuracy, improving error detection rates
by 8%, and reducing QC errors by 10%, enhancing client satisfaction.

Credit Card Fraud Detection: Developed a high-performance credit card fraud detection model using
Python and Random Forest classifier, resulting in a 10% reduction in false positive cases within the first
quarter of deployment. The model achieved 96% precision, 92% recall, and an AUC of 0.97.

Data Architecture & Insights: Performed end-to-end data exploration, transformation, and feature
engineering on complex call center datasets, optimizing KPI metrics such as AHT and NPS for global
banking clients.

Customer Segmentation: Engineered a scalable LightGBM predictive model to identify high-potential
telecom customers, doubling Click-Through Rate (CTR), improving profit margins by 50%, and reducing
Cost Per Lead (CPL) by 20%.

Sentiment Analysis & NLP: Leveraged pre-trained BERT models to analyze customer feedback,
extracting quantifiable sentiment scores and satisfaction drivers for actionable insights.

Lead Scoring Models: Implemented XGBoost models for sales lead prioritisation, improving lead
conversion prediction accuracy by 20% and reducing sales cycle length by 5%.

"""

SENIOR_YEARS = 10  # Apoorv's years of experience


def extract_years_required(description: str) -> int | None:
    """Best-effort extraction of minimum years-of-experience requirement."""
    matches = re.findall(r"(\d+)\+?\s*(?:years|yrs|jahre)", description, re.IGNORECASE)
    if not matches:
        return None
    return min(int(m) for m in matches)


def detect_german_requirement(description: str) -> str:
    """Classify German language requirement from job description text."""
    text = description.lower()

    strong_german_markers = [
        "verhandlungssicher", "fließend deutsch", "sehr gute deutschkenntnisse",
        "c1-niveau", "c1 niveau", "c1-c2", "german c1", "fluent german",
        "native german", "muttersprache deutsch",
    ]
    mentions_german = "deutsch" in text or "german" in text
    mentions_english_sufficient = (
        "english fluency" in text or "fluent in english" in text
        or "business english" in text or "professional english" in text
    ) and "german" not in text

    if any(marker in text for marker in strong_german_markers):
        return "Yes - fluent/C1+ German explicitly required"
    if mentions_german and "plus" in text and "german" in text.split("plus")[0][-60:]:
        return "No - German mentioned only as a nice-to-have"
    if mentions_german:
        return "Yes - German required (see JD for exact level)"
    if mentions_english_sufficient:
        return "No - English fluency stated as sufficient"
    return "Not stated in JD"


def detect_sponsorship(description: str) -> str:
    text = description.lower()
    if "visa sponsorship" in text or "we offer visa" in text or "sponsor your visa" in text:
        return "Yes - explicitly stated in JD"
    if "relocation" in text and ("visa" in text or "work permit" in text):
        return "Yes - relocation & visa support mentioned in JD"
    if "relocation" in text:
        return "Possible - relocation support mentioned, visa not explicit"
    return "Not stated in JD"


# ---------------------------------------------------------------------------
# Gemini-based scoring
# ---------------------------------------------------------------------------

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
GEMINI_MODEL = "gemini-flash-lite-latest"


class JobFitScore(BaseModel):
    score: int = Field(
        ge=0, le=100,
        description="Fit score 0-100 for how well this job matches the candidate profile.",
    )
    reason: str = Field(
        description=(
            "One or two concise sentences explaining the score: what matches "
            "the candidate's GenAI/LLM/agentic specialization specifically, "
            "what doesn't, and any seniority mismatch."
        ),
    )


_gemini_llm = None


def _get_gemini_llm():
    global _gemini_llm
    if _gemini_llm is None:
        _gemini_llm = ChatGoogleGenerativeAI(
            model=GEMINI_MODEL,
            google_api_key=GEMINI_API_KEY,
            temperature=0,
        ).with_structured_output(JobFitScore)
    return _gemini_llm

SCORING_INSTRUCTIONS = """
You are screening a job posting against a specific candidate's resume to produce an ATS-style
match score from 0 to 100.

Scoring guidance:
- Weight the candidate's core specialization (GenAI, LLMs, RAG, LangChain, LangGraph, agentic AI,
  Azure OpenAI, prompt engineering) heavily. A role built around these should score high (80-100)
  if seniority also lines up.
- A role that only touches classical ML/statistics with GenAI as a passing mention ("exposure to
  LLMs is a plus") is a meaningfully weaker fit than a true GenAI-specialist role, even if it
  shares some keywords -- score it lower (roughly 25-45), not just slightly lower.
- A role clearly unrelated to data science/AI entirely should score very low (under 15).
- Adjust for seniority: a role requiring far more years/seniority than the candidate has (e.g.
  15+ years, VP/Director/Chief-level title well above the candidate's level) or far less (pure
  internship/Werkstudent/fresh-graduate program) is a weaker fit -- reflect this in the score,
  not just the reason text.
- Do not penalize a role for language requirements, sponsorship, or location -- those are tracked
  separately. Score purely on technical/domain/seniority fit.
- Be honest and calibrated, not generous. A mediocre fit should score like a mediocre fit.

Respond only via the provided JSON schema.
"""


def score_job_with_gemini(job: dict) -> tuple[int, str]:
    """Score a job against RESUME_PROFILE using the Gemini API. Raises on failure
    (caller is responsible for falling back)."""
    title = job.get("title", "")
    company = job.get("companyName", "")
    description = (job.get("descriptionText") or "")[:6000]  # bound prompt size

    prompt = f"""{SCORING_INSTRUCTIONS}

CANDIDATE PROFILE:
{RESUME_PROFILE}

CANDIDATE WORK EXPERIENCE:
{WORK_EXPERIENCE}

JOB POSTING:
Title: {title}
Company: {company}
Description:
{description}
"""

    llm = _get_gemini_llm()
    result = llm.invoke(prompt)
    if not isinstance(result, JobFitScore):
        raise ValueError(f"Unexpected Gemini response shape: {result!r}")
    return max(0, min(100, result.score)), result.reason


def score_job_fallback(job: dict) -> tuple[int, str]:
    """Deterministic keyword-based fallback, used only if the Gemini call fails
    (network issue, quota exhausted, malformed response, etc.) so a single API
    hiccup doesn't stop the whole run or silently drop a job."""
    description = (job.get("descriptionText") or "").lower()
    title = (job.get("title") or "").lower()
    combined_text = f"{description} {title}"

    core_terms = ["genai", "generative ai", "rag", "langchain", "langgraph", "agentic", "azure openai"]
    hits = sum(1 for t in core_terms if t in combined_text)
    score = min(90, hits * 15) if hits else 10
    reason = (
        f"[FALLBACK SCORER - Gemini call failed] Matched {hits} core GenAI term(s) "
        f"via keyword search. Verify manually; this is a degraded-mode score."
    )
    return score, reason


_gemini_quota_exhausted = False


def score_job(job: dict) -> tuple[int, str, bool]:
    """Returns (score 0-100, reason string, scored_by_gemini). Tries Gemini first,
    falls back to a deterministic keyword scorer on any failure so the run doesn't
    crash. `scored_by_gemini` tells the caller whether it's safe to mark the job
    permanently "seen" -- fallback-scored jobs are left unmarked so a later run
    (once quota resets) retries them for a real score instead of being stuck."""
    global _gemini_quota_exhausted

    if not GEMINI_API_KEY:
        print("WARNING: GEMINI_API_KEY not set -- using fallback keyword scorer for all jobs.", file=sys.stderr)
        return (*score_job_fallback(job), False)

    if _gemini_quota_exhausted:
        # Already hit the daily quota this run -- don't waste time/requests
        # re-hitting it for every remaining job, just fall back straight away.
        return (*score_job_fallback(job), False)

    try:
        score, reason = score_job_with_gemini(job)
        return score, reason, True
    except ModelRateLimitError as e:
        _gemini_quota_exhausted = True
        print(f"WARNING: Gemini quota exhausted ({e}). Falling back to keyword scorer for the rest of this run.", file=sys.stderr)
        return (*score_job_fallback(job), False)
    except Exception as e:
        print(f"WARNING: Gemini scoring failed for '{job.get('title')}': {e}. Using fallback scorer.", file=sys.stderr)
        return (*score_job_fallback(job), False)


# ---------------------------------------------------------------------------
# Apify fetch
# ---------------------------------------------------------------------------

def fetch_jobs(keywords: str, location: str, limit: int = 30) -> list[dict]:
    payload = {
        "keywords": keywords,
        "location": location,
        "datePosted": "past24Hours",
        "limitPerSource": limit,
        "scrapeCompany": False,
    }
    resp = requests.post(
        APIFY_RUN_URL,
        params={"token": APIFY_TOKEN},
        json=payload,
        timeout=180,
    )
    resp.raise_for_status()
    return resp.json()


def job_uid(job: dict) -> str:
    """Stable unique ID for a job posting, used for dedup across runs."""
    job_id = job.get("id") or job.get("trackingId")
    if job_id:
        return str(job_id)
    # fallback: derive from link, stripped of query params
    link = job.get("link", "")
    return link.split("?")[0]


# ---------------------------------------------------------------------------
# Persistence (seen jobs across runs)
# ---------------------------------------------------------------------------

def load_seen_jobs() -> dict:
    if SEEN_JOBS_PATH.exists():
        with open(SEEN_JOBS_PATH, "r") as f:
            return json.load(f)
    return {}


def save_seen_jobs(seen: dict) -> None:
    with open(SEEN_JOBS_PATH, "w") as f:
        json.dump(seen, f, indent=2, sort_keys=True)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

FONT_NAME = "Arial"
HEADERS = [
    "Rank", "Title", "Company", "Location", "Date Posted",
    "Match Score (%)", "ATS Status (>75%)", "Reason",
    "German Fluency Required", "Sponsorship Provided",
    "Fetch Run", "Still Open", "Direct Link",
]


def style_header(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_row(ws, row_idx: int, rank: int, rec: dict):
    border = Border(*[Side(style="thin", color="B7B7B7")] * 4)
    qualify_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    below_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    new_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    closed_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    status = "APPLY" if rec["score"] >= ATS_THRESHOLD else "Below threshold"
    values = [
        rank, rec["title"], rec["company"], rec["location"], rec["date_posted"],
        rec["score"], status, rec["reason"], rec["german_required"],
        rec["sponsorship"], rec["fetch_run"], rec["still_open"], rec["link"],
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.border = border
        cell.alignment = Alignment(
            wrap_text=True, vertical="top",
            horizontal="center" if col_idx in (1, 5, 6, 7, 11, 12) else "left",
        )

    row_fill = qualify_fill if rec["score"] >= ATS_THRESHOLD else below_fill
    ws.cell(row=row_idx, column=6).fill = row_fill
    ws.cell(row=row_idx, column=7).fill = row_fill
    ws.cell(row=row_idx, column=7).font = Font(name=FONT_NAME, size=10, bold=True)

    if rec["fetch_run"].startswith("New"):
        ws.cell(row=row_idx, column=11).fill = new_fill
        ws.cell(row=row_idx, column=11).font = Font(name=FONT_NAME, size=10, bold=True)

    if rec["still_open"] == "Closed":
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = closed_fill

    link_cell = ws.cell(row=row_idx, column=13)
    link_cell.hyperlink = rec["link"]
    link_cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")


def rebuild_workbook(all_records: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    ws.append(HEADERS)
    style_header(ws)

    # Sort: open jobs by score desc first, then closed jobs at the bottom
    open_jobs = sorted(
        [r for r in all_records if r["still_open"] != "Closed"],
        key=lambda r: -r["score"],
    )
    closed_jobs = [r for r in all_records if r["still_open"] == "Closed"]
    ordered = open_jobs + closed_jobs

    for idx, rec in enumerate(ordered, start=1):
        write_row(ws, idx + 1, idx, rec)

    widths = {
        "A": 6, "B": 40, "C": 24, "D": 20, "E": 12, "F": 12, "G": 15,
        "H": 50, "I": 28, "J": 34, "K": 20, "L": 12, "M": 38,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    for r in range(2, len(ordered) + 2):
        ws.row_dimensions[r].height = 55

    wb.save(OUTPUT_XLSX)
    print(f"Saved {len(ordered)} rows to {OUTPUT_XLSX}")


def load_existing_records() -> dict:
    """Load existing rows from the workbook (keyed by job link) if present."""
    if not OUTPUT_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(OUTPUT_XLSX)
    ws = wb["Job Matches"]
    records = {}
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        link = row[idx["Direct Link"]]
        if not link:
            continue
        records[link] = {
            "title": row[idx["Title"]],
            "company": row[idx["Company"]],
            "location": row[idx["Location"]],
            "date_posted": row[idx["Date Posted"]],
            "score": row[idx["Match Score (%)"]],
            "reason": row[idx["Reason"]],
            "german_required": row[idx["German Fluency Required"]],
            "sponsorship": row[idx["Sponsorship Provided"]],
            "fetch_run": "From previous run",  # downgrade any old "New" tag
            "still_open": row[idx["Still Open"]] or "Open",
            "link": link,
        }
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    run_timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    seen = load_seen_jobs()
    existing_records = load_existing_records()

    new_count = 0
    new_jobs = []
    all_raw_jobs = []

    for search in SEARCHES:
        try:
            jobs = fetch_jobs(search["keywords"], search["location"])
        except requests.RequestException as e:
            print(f"WARNING: fetch failed for {search}: {e}", file=sys.stderr)
            continue
        all_raw_jobs.extend(jobs)
        print(f"Fetched {len(jobs)} raw results for '{search['keywords']}'")

    for job in all_raw_jobs:
        if is_excluded(job):
            continue

        uid = job_uid(job)
        link = job.get("link", "")

        if uid in seen:
            continue  # already tracked in a previous run, skip entirely

        score, reason, scored_by_gemini = score_job(job)
        description = job.get("descriptionText") or ""

        record = {
            "title": job.get("title", ""),
            "company": job.get("companyName", ""),
            "location": job.get("location", ""),
            "date_posted": job.get("postedAt", ""),
            "score": score,
            "reason": reason,
            "german_required": detect_german_requirement(description),
            "sponsorship": detect_sponsorship(description),
            "fetch_run": f"New this run ({run_timestamp})",
            "still_open": "Open",
            "link": link,
        }
        existing_records[link] = record
        if scored_by_gemini:
            seen[uid] = {"first_seen": run_timestamp, "link": link}
        # else: leave unmarked so a later run (once quota resets) retries
        # this job for a real Gemini score instead of it being stuck on the
        # degraded fallback score forever.
        new_count += 1
        new_jobs.append(record)

    print(f"{new_count} new job(s) added this run.")

    # Closure re-check disabled: notifications now go via Telegram instead of
    # the spreadsheet, and there's no persisted record store left to re-check
    # previously-tracked listings against.

    # rebuild_workbook(list(existing_records.values()))  # disabled: notifications now go via Telegram instead of the spreadsheet
    save_seen_jobs(seen)

    github_output = os.environ.get("GITHUB_OUTPUT")
    if github_output:
        # Only notify about jobs that cleared the match-score bar -- a new job
        # below threshold is still tracked/deduped, just not worth pinging about.
        notify_jobs = [j for j in new_jobs if j["score"] > ATS_THRESHOLD]

        # Telegram hard-caps messages at 4096 UTF-8 chars. Budget for the summary
        # line (~100 chars) and build the job list up to that budget, cutting off
        # by size rather than a fixed job count since titles/links vary in length.
        MAX_MESSAGE_CHARS = 4096
        JOB_LIST_BUDGET = MAX_MESSAGE_CHARS - 200
        lines = []
        included = 0
        for j in notify_jobs:
            entry = f"{j['title']} @ {j['company']} (score: {j['score']})\n{j['link']}"
            projected_len = len("\n\n".join(lines + [entry]))
            if projected_len > JOB_LIST_BUDGET:
                break
            lines.append(entry)
            included += 1
        if included < len(notify_jobs):
            lines.append(f"... and {len(notify_jobs) - included} more.")
        new_jobs_text = "\n\n".join(lines)

        with open(github_output, "a") as f:
            f.write(f"new_count={len(notify_jobs)}\n")
            f.write("new_jobs<<GH_OUTPUT_EOF\n")
            f.write(f"{new_jobs_text}\n")
            f.write("GH_OUTPUT_EOF\n")


if __name__ == "__main__":
    main()
